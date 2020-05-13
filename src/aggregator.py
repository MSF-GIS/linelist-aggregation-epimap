import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
import Levenshtein


def get_data_dir():
    """
    Get the path of the data directory
    :return:
    """
    if os.path.exists('data'):
        return 'data'
    elif os.path.join('..', 'data'):
        return os.path.join('..', 'data')
    raise Exception('Data directory not found')


def fix_village_if_necessary(linelist, villages):
    idx = linelist[~linelist['Village'].isin(villages)].index
    for i in idx:
        min_distance = 5
        min_village = ''
        for village in villages:
            dist = Levenshtein.distance(linelist.loc[i, 'Village'], village)
            if dist < min_distance:
                min_distance = dist
                min_village = village
        if min_village == '':
            logging.error(f'Line {i} of Linelist removed because village not found : ' + linelist.loc[i, 'Village'])
            linelist = linelist[linelist.index != i]
        else:
            linelist.loc[i, 'Village'] = min_village
    return linelist


def aggregate(input_file):
    """
    Read input file and aggregate the number of case by week and village

    :param input_file: The Linelist Excel file
    :return: Dataframe whit aggregation. Columns : 'CODE_LOCATION', 'Year', 'Month', 'Week', 'Num of cases'
    """
    # Get data
    linelist = pd.read_excel(input_file, sheet_name='Linelist')[['Date of Assessment', 'Village']]
    linelist = linelist[~pd.isnull(linelist).transpose().any()]
    epiweek = pd.read_excel(input_file, sheet_name='Epiweeks', header=2)[['Epi week', 'Month', 'First day in week']]
    geo = pd.read_excel(input_file, sheet_name='GEO')[['Camp/village', 'GEOID (pcode)?']]
    geo.columns = ['Village', 'CODE_LOCATION']

    # Transform data
    linelist['Date of Assessment'] = pd.to_datetime(linelist['Date of Assessment'])
    linelist['Week'] = linelist['Date of Assessment'].apply(lambda t: t.isocalendar()[1])
    linelist['Year'] = linelist['Date of Assessment'].apply(lambda t: t.year)
    linelist = fix_village_if_necessary(linelist, geo['Village'].unique())
    linelist_agg = linelist.groupby(['Village', 'Week', 'Year']).count()
    linelist_agg.columns = ['Num of cases']
    linelist_agg = linelist_agg.reset_index()
    epiweek['Year'] = -1
    for i in range(len(epiweek)):
        if type(epiweek.loc[i, 'First day in week']) is str:
            epiweek.loc[i, 'First day in week'] = datetime.datetime.strptime(epiweek.loc[i, 'First day in week'], '%d/%m/%Y')
        epiweek.loc[i, 'Year'] = epiweek.loc[i, 'First day in week'].year
        if epiweek.loc[i, 'Epi week'] == 1 and epiweek.loc[i, 'First day in week'].month == 12:
            epiweek.loc[i, 'Year'] += 1
    epiweek = epiweek[['Epi week', 'Month', 'Year']]
    epiweek.columns = ['Week', 'Month', 'Year']

    # Merge data
    linelist_agg_geo = pd.merge(linelist_agg, geo, how='left', on='Village')
    res = pd.DataFrame()
    villages = geo['CODE_LOCATION'].unique()
    villages = villages[~pd.isnull(villages)]
    for code_loc in sorted(villages):
        code_loc_linelist = linelist_agg_geo[linelist_agg_geo['CODE_LOCATION'] == code_loc]
        code_loc_cases = pd.merge(epiweek, code_loc_linelist, how='left', on=['Year', 'Week'])
        code_loc_cases['CODE_LOCATION'] = code_loc
        res = pd.concat([res, code_loc_cases], axis=0, ignore_index=True)
    return res[['CODE_LOCATION', 'Year', 'Month', 'Week', 'Num of cases']]


def export_data_frame_to_excel(df, output_file):
    """
    Export a dataframe in Excel with style

    :param df: The Dataframe which will be exported
    :param output_file: The output Excel file
    :return: None
    """
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Apply color and border
    thin = Side(border_style="thin", color="000000")
    for col in ['A', 'B', 'C', 'D']:
        for index in range(1, len(df) + 2):
            ws[col + str(index)].fill = PatternFill("solid", fgColor="D0CECE")
            ws[col + str(index)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for index in range(1, len(df) + 2):
        ws['E' + str(index)].fill = PatternFill("solid", fgColor="C6E0B4")
        ws['E' + str(index)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # Apply bold
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[col + '1'].font = Font(bold=True)
    for index in range(1, len(df) + 2):
        ws['A' + str(index)].font = Font(bold=True)
    # Apply column width
    ws.column_dimensions['A'].width = 21.71
    ws.column_dimensions['B'].width = 8.43
    ws.column_dimensions['C'].width = 14.57
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 26.57
    wb.save(output_file)


if __name__ == '__main__':
    logging.basicConfig(filename='logs.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    input_file_name = 'Linelist example_input file.xlsx'
    try:
        df = aggregate(os.path.join(get_data_dir(), input_file_name))
        export_data_frame_to_excel(df, os.path.join(get_data_dir(), 'res_V3.xlsx'))
    except Exception as e:
        logging.exception('Error: ' + str(e))
